from openpyxl import *
import numpy as np
from matplotlib.mlab import bivariate_normal

import scipy.stats as st
import math
# Change the directory to where you store your data
work_book = load_workbook("/Users/jingyizhao/PycharmProjects/TensorFlow /university data.xlsx")
# load sheet
work_sheet = work_book['university_data']
rank = []
name = []
CS_score = []
Research_Overhead = []
Admin_Base_Pay = []
Tuition = []
Grad_Student_No = []
for i in range(2, 51):
    rank.append(work_sheet.cell(row=i, column=1).value)
    name.append(work_sheet.cell(row=i, column=2).value)
    CS_score.append(work_sheet.cell(row=i, column=3).value)
    Research_Overhead.append(work_sheet.cell(row=i, column=4).value)
    Admin_Base_Pay.append(work_sheet.cell(row=i, column=5).value)
    Tuition.append(work_sheet.cell(row=i, column=6).value)
    Grad_Student_No.append(work_sheet.cell(row=i, column=7).value)

mu1 = round(np.mean(CS_score), 3)  # average of CS_score
mu2 = round(np.mean(Research_Overhead), 3)  # average of Research_Overhead
mu3 = round(np.mean(Admin_Base_Pay), 3)  # average of Admin_Base_Pay
mu4 = round(np.mean(Tuition), 3)  # average of Tuition

var1 = round(np.var(CS_score), 3)  # Variance of Cs_score
var2 = round(np.var(Research_Overhead), 3)  # Variance of Research_Overhead
var3 = round(np.var(Admin_Base_Pay), 3)  # Variance of Admin_Base_Pay
var4 = round(np.var(Tuition), 3)  # Variance of Tuition

sigma1 = round(np.std(CS_score), 3)  # Standard deviation of Cs_score
sigma2 = round(np.std(Research_Overhead), 3)  # Standard deviation of Research_Overhead
sigma3 = round(np.std(Admin_Base_Pay), 3)  # Standard deviation of Admin_Base_Pay
sigma4 = round(np.std(Tuition), 3)  # Standard deviation of Tuition

# covarianceMat & CorrelationMat
y = [CS_score, Research_Overhead, Admin_Base_Pay, Tuition]
covarianceMat = np.cov(y)
correlationMat = np.corrcoef(y)

# Loglikehood
px1 = st.norm.pdf(CS_score, loc=mu1, scale=var1)
px2 = st.norm.pdf(Research_Overhead, loc=mu2, scale=var2)
px3 = st.norm.pdf(Admin_Base_Pay, loc=mu3, scale=var3)
px4 = st.norm.pdf(Tuition, loc=mu4, scale=var4)











def Loglikehood(px1,px2,px3,px4):
    sum = 0;
    for i in range(0, 48):
        p = px1[i]*px2[i]*px3[i]*px4[i];
        sum +=  math.log(p)
    return sum


# multivariate Loglikehood

# def MutiLoglikehood(px):
#     means = [mu1, mu2, mu3, mu4]
#     y_trans = np.transpose(y)
#     muti_sum = 0
#     for example in y_trans:
#         example = math.log(st.multivariate_normal.pdf(example, mean=means, cov=covarianceMat))
#         muti_sum = + example


def pdf_multivariate_gauss(x, mu, cov):
    '''
    Caculate the multivariate normal density (pdf)
    Keyword arguments:
        x = numpy array of a "d x 1" sample vector
        mu = numpy array of a "d x 1" mean vector
        cov = "numpy array of a d x d" covariance matrix
    '''
    part1 = 1 / (((2 * np.pi) ** (len(mu) / 2)) * (np.linalg.det(cov) ** (1 / 2)))
    part2 = (-1 / 2) * ((x - mu).T.dot(np.linalg.inv(cov))).dot((x - mu))
    return float(part1 * np.exp(part2))


def test_gauss_pdf():

    multi_sum = 0;
    for i in range(0,48):
        x = np.array([CS_score[i],Research_Overhead[i],Admin_Base_Pay[i],Tuition[i]])
        mu =np.array([mu1,mu2,mu3,mu4]);
        cov = covarianceMat;
        p = pdf_multivariate_gauss(x, mu, cov);
        multi_sum += math.log(p);

    print('multi_variable model logP', multi_sum)





if __name__ == '__main__':
    test_gauss_pdf()
    sum = Loglikehood(px1,px2,px3,px4);
    print("Independent Variable Model LogP", sum);

# #
# print('mu1 = ', mu1)
# print('mu2 = ', mu2)
# print('mu3 = ', mu3)
# print('mu4 = ', mu4)
# print('var1=', var1)
# print('var2=', var2)
# print('var3=', var3)
# print('var4=', var4)
# print('sigma1=', sigma1)
# print('sigma2=', sigma2)
# print('sigma3=', sigma3)
# print('sigma4=', sigma4)



# print('covarianceMat=\n', covarianceMat)
# print('correlationMat=\n', correlationMat)



